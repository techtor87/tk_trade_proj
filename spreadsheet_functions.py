#!/usr/bin/env python

import numpy as np
import openpyxl as pyxl
import string

from tk_functions import *
from trading_stratigies import *

def setup_columns( _worksheet ):
    _worksheet.cell(row=1, column=1, value='Cur Bid')
    _worksheet.cell(row=1, column=2, value='Cur Ask')
    _worksheet.cell(row=1, column=3, value='strikeprice')
    _worksheet.cell(row=1, column=4, value='xdate')
    _worksheet.cell(row=1, column=5, value='days_to_expir')

    _worksheet.cell(row=1, column=7, value='bid')
    _worksheet.cell(row=1, column=8, value='ask')
    _worksheet.cell(row=1, column=9, value='contract_size')
    _worksheet.cell(row=1, column=10, value='p/c')
    _worksheet.cell(row=1, column=11, value='symbol')
    _worksheet.cell(row=1, column=12, value='imp_vol')
    _worksheet.cell(row=1, column=13, value='delta')
    _worksheet.cell(row=1, column=14, value='gamma')
    _worksheet.cell(row=1, column=15, value='theta')
    _worksheet.cell(row=1, column=16, value='vega')
    _worksheet.cell(row=1, column=17, value='rho')
    _worksheet.cell(row=1, column=18, value='openinterest')

    _worksheet.cell(row=1, column=20, value='bid')
    _worksheet.cell(row=1, column=21, value='ask')
    _worksheet.cell(row=1, column=22, value='contract_size')
    _worksheet.cell(row=1, column=23, value='p/c')
    _worksheet.cell(row=1, column=24, value='symbol')
    _worksheet.cell(row=1, column=25, value='imp_vol')
    _worksheet.cell(row=1, column=26, value='delta')
    _worksheet.cell(row=1, column=27, value='gamma')
    _worksheet.cell(row=1, column=28, value='theta')
    _worksheet.cell(row=1, column=29, value='vega')
    _worksheet.cell(row=1, column=30, value='rho')
    _worksheet.cell(row=1, column=31, value='openinterest')
    return _worksheet


def setup_trade_columns( _tradesheet ):
    _tradesheet.cell(row=1, column=1, value='Trade Type')
    _tradesheet.cell(row=1, column=2, value='Bid')
    _tradesheet.cell(row=1, column=3, value='Ask')
    _tradesheet.cell(row=1, column=4, value='Trade Profit')
    last_column = 5
    for i in xrange(1,5):
        _tradesheet.cell(row=1, column=last_column+1, value='put_call')
        _tradesheet.cell(row=1, column=last_column+2, value='rootsymbol')
        _tradesheet.cell(row=1, column=last_column+3, value='strikeprice')
        _tradesheet.cell(row=1, column=last_column+4, value='xdate')
        _tradesheet.cell(row=1, column=last_column+5, value='ask')
        _tradesheet.cell(row=1, column=last_column+6, value='ask_time')
        _tradesheet.cell(row=1, column=last_column+7, value='asksz')
        _tradesheet.cell(row=1, column=last_column+8, value='basis')
        _tradesheet.cell(row=1, column=last_column+9, value='bid')
        _tradesheet.cell(row=1, column=last_column+10, value='bid_time')
        _tradesheet.cell(row=1, column=last_column+11, value='bidsz')
        _tradesheet.cell(row=1, column=last_column+12, value='chg')
        _tradesheet.cell(row=1, column=last_column+13, value='chg_sign')
        _tradesheet.cell(row=1, column=last_column+14, value='chg_t')
        _tradesheet.cell(row=1, column=last_column+15, value='cl')
        _tradesheet.cell(row=1, column=last_column+16, value='days_to_expr')
        _tradesheet.cell(row=1, column=last_column+17, value='exch')
        _tradesheet.cell(row=1, column=last_column+18, value='exch_desc')
        _tradesheet.cell(row=1, column=last_column+19, value='hi')
        _tradesheet.cell(row=1, column=last_column+20, value='incr_vl')
        _tradesheet.cell(row=1, column=last_column+21, value='issue_desc')
        _tradesheet.cell(row=1, column=last_column+22, value='last')
        _tradesheet.cell(row=1, column=last_column+23, value='lo')
        _tradesheet.cell(row=1, column=last_column+24, value='op_delivery')
        _tradesheet.cell(row=1, column=last_column+25, value='op_flag')
        _tradesheet.cell(row=1, column=last_column+26, value='op_style')
        _tradesheet.cell(row=1, column=last_column+27, value='op_subclass')
        _tradesheet.cell(row=1, column=last_column+28, value='opn')
        _tradesheet.cell(row=1, column=last_column+29, value='pchg')
        _tradesheet.cell(row=1, column=last_column+30, value='pchg_sign')
        _tradesheet.cell(row=1, column=last_column+31, value='pcls')
        _tradesheet.cell(row=1, column=last_column+32, value='phi')
        _tradesheet.cell(row=1, column=last_column+33, value='plo')
        _tradesheet.cell(row=1, column=last_column+34, value='popn')
        _tradesheet.cell(row=1, column=last_column+35, value='pr_date')
        _tradesheet.cell(row=1, column=last_column+36, value='pr_openinterest')
        _tradesheet.cell(row=1, column=last_column+37, value='prchg')
        _tradesheet.cell(row=1, column=last_column+38, value='prem_mult')
        _tradesheet.cell(row=1, column=last_column+39, value='pvol')
        _tradesheet.cell(row=1, column=last_column+40, value='secclass')
        _tradesheet.cell(row=1, column=last_column+41, value='sesn')
        _tradesheet.cell(row=1, column=last_column+42, value='symbol')
        _tradesheet.cell(row=1, column=last_column+43, value='tcond')
        _tradesheet.cell(row=1, column=last_column+44, value='timestamp')
        _tradesheet.cell(row=1, column=last_column+45, value='tr_num')
        _tradesheet.cell(row=1, column=last_column+46, value='tradetick')
        _tradesheet.cell(row=1, column=last_column+47, value='under_susip')
        _tradesheet.cell(row=1, column=last_column+48, value='vl')
        _tradesheet.cell(row=1, column=last_column+49, value='vwap')
        _tradesheet.cell(row=1, column=last_column+50, value='wk52hi')
        _tradesheet.cell(row=1, column=last_column+51, value='wk52hidate')
        _tradesheet.cell(row=1, column=last_column+52, value='wk52lo')
        _tradesheet.cell(row=1, column=last_column+53, value='wk52lodate')
        _tradesheet.cell(row=1, column=last_column+54, value='imp_vol')
        _tradesheet.cell(row=1, column=last_column+55, value='idelta')
        _tradesheet.cell(row=1, column=last_column+56, value='igamma')
        _tradesheet.cell(row=1, column=last_column+57, value='itheta')
        _tradesheet.cell(row=1, column=last_column+58, value='ivega')
        _tradesheet.cell(row=1, column=last_column+59, value='irho')
        _tradesheet.cell(row=1, column=last_column+60, value='openinterest')

        last_column = last_column + 61

    return

def fill_row( _worksheet, _data, _bid, _ask ):
    # new_row_index = _worksheet.row
    # if( ! _data.search_quote.empty? )
    #     for quote in _data.search_quote.sort_by { |a| [a.xdate, a.strikeprice, a.put_call ] }
    #         if quote.put_call == "call"
    #         then
    #             if ( ( _worksheet[new_row_index, 2] != quote.strikeprice ) )
    #             then
    #                 new_row_index += 1

    #                 _worksheet[new_row_index, 0] = _bid
    #                 _worksheet[new_row_index, 1] = _ask
    #                 _worksheet[new_row_index, 2] = quote.strikeprice
    #                 _worksheet[new_row_index, 3] = quote.xdate
    #                 _worksheet[new_row_index, 4] = quote.days_to_expir
                # Call
                # _worksheet[new_row_index,  6] = quote.bid
                # _worksheet[new_row_index,  7] = quote.ask
                # _worksheet[new_row_index,  8] = quote.contract_size
                # _worksheet[new_row_index,  9] = quote.put_call
                # _worksheet[new_row_index,  10] = quote.symbol
                # _worksheet[new_row_index,  11] = quote.imp_volatility
                # _worksheet[new_row_index,  12] = quote.idelta
                # _worksheet[new_row_index,  13] = quote.igamma
                # _worksheet[new_row_index,  14] = quote.itheta
                # _worksheet[new_row_index,  15] = quote.ivega
                # _worksheet[new_row_index,  16] = quote.irho
                # _worksheet[new_row_index,  17] = quote.openinterest
            # elsif quote.put_call == "put"
            # then
                # if ( ( _worksheet[new_row_index, 2] != quote.strikeprice ) ) # && !_worksheet[new_row_index, 2].to_s.empty?
                # then
                    # new_row_index += 1

                    # _worksheet[new_row_index,  0] = _bid
                    # _worksheet[new_row_index,  1] = _ask
                    # _worksheet[new_row_index,  2] = quote.strikeprice
                    # _worksheet[new_row_index,  3] = quote.xdate
                    # _worksheet[new_row_index,  4] = quote.days_to_expiration

                # Puts
                # _worksheet[new_row_index,  6] = quote.bid
                # _worksheet[new_row_index,  7] = quote.ask
                # _worksheet[new_row_index,  8] = quote.contract_size
                # _worksheet[new_row_index,  9] = quote.put_call
                # _worksheet[new_row_index,  10] = quote.symbol
                # _worksheet[new_row_index,  11] = quote.imp_volatility
                # _worksheet[new_row_index,  12] = quote.idelta
                # _worksheet[new_row_index,  13] = quote.igamma
                # _worksheet[new_row_index,  14] = quote.itheta
                # _worksheet[new_row_index,  15] = quote.ivega
                # _worksheet[new_row_index,  16] = quote.irho
                # _worksheet[new_row_index,  17] = quote.openinterest

                # new_row_index += 1
    return

def print_trades( _tradesheet, _bid, _ask, _type, _profit, *_opts ):
         # new_row_index = _tradesheet.last_row_index + 1
         # _tradesheet[new_row_index, 0] = _type
         # _tradesheet[new_row_index, 1] = _bid
         # _tradesheet[new_row_index, 2] = _ask
         # _tradesheet[new_row_index, 3] = _profit
         # last_column = 4
         # for opt in _opts
         # _tradesheet[new_row_index, last_column+1] = opt.put_call
         # _tradesheet[new_row_index, last_column+2] = opt.rootsymbol
         # _tradesheet[new_row_index, last_column+3] = opt.strikeprice
         # _tradesheet[new_row_index, last_column+4] = opt.xdate
         # _tradesheet[new_row_index, last_column+5] = opt.ask
         # _tradesheet[new_row_index, last_column+6] = opt.ask_time
         # _tradesheet[new_row_index, last_column+7] = opt.asksz
         # _tradesheet[new_row_index, last_column+8] = opt.basis
         # _tradesheet[new_row_index, last_column+9] = opt.bid
         # _tradesheet[new_row_index, last_column+10] = opt.bid_time
         # _tradesheet[new_row_index, last_column+11] = opt.bidsz
         # _tradesheet[new_row_index, last_column+12] = opt.chg
         # _tradesheet[new_row_index, last_column+13] = opt.chg_sign
         # _tradesheet[new_row_index, last_column+14] = opt.chg_t
         # _tradesheet[new_row_index, last_column+15] = opt.cl
         # _tradesheet[new_row_index, last_column+16] = opt.days_to_expiration
         # _tradesheet[new_row_index, last_column+17] = opt.exch
         # _tradesheet[new_row_index, last_column+18] = opt.exch_desc
         # _tradesheet[new_row_index, last_column+19] = opt.hi
         # _tradesheet[new_row_index, last_column+20] = opt.incr_vl
         # _tradesheet[new_row_index, last_column+21] = opt.issue_desc
         # _tradesheet[new_row_index, last_column+22] = opt.last
         # _tradesheet[new_row_index, last_column+23] = opt.lo
         # _tradesheet[new_row_index, last_column+24] = opt.op_delivery
         # _tradesheet[new_row_index, last_column+25] = opt.op_flag
         # _tradesheet[new_row_index, last_column+26] = opt.op_style
         # _tradesheet[new_row_index, last_column+27] = opt.op_subclass
         # _tradesheet[new_row_index, last_column+28] = opt.opn
         # _tradesheet[new_row_index, last_column+29] = opt.pchg
         # _tradesheet[new_row_index, last_column+30] = opt.pchg_sign
         # _tradesheet[new_row_index, last_column+31] = opt.pcls
         # _tradesheet[new_row_index, last_column+32] = opt.phi
         # _tradesheet[new_row_index, last_column+33] = opt.plo
         # _tradesheet[new_row_index, last_column+34] = opt.popn
         # _tradesheet[new_row_index, last_column+35] = opt.pr_date
         # _tradesheet[new_row_index, last_column+36] = opt.pr_openinterest
         # _tradesheet[new_row_index, last_column+37] = opt.prchg
         # _tradesheet[new_row_index, last_column+38] = opt.prem_mult
         # _tradesheet[new_row_index, last_column+39] = opt.pvol
         # _tradesheet[new_row_index, last_column+40] = opt.secclass
         # _tradesheet[new_row_index, last_column+41] = opt.sesn
         # _tradesheet[new_row_index, last_column+42] = opt.symbol
         # _tradesheet[new_row_index, last_column+43] = opt.tcond
         # _tradesheet[new_row_index, last_column+44] = opt.timestamp
         # _tradesheet[new_row_index, last_column+45] = opt.tr_num
         # _tradesheet[new_row_index, last_column+46] = opt.tradetick
         # _tradesheet[new_row_index, last_column+47] = opt.under_cusip
         # _tradesheet[new_row_index, last_column+48] = opt.vl
         # _tradesheet[new_row_index, last_column+49] = opt.vwap
         # _tradesheet[new_row_index, last_column+50] = opt.wk52hi
         # _tradesheet[new_row_index, last_column+51] = opt.wk52hidate
         # _tradesheet[new_row_index, last_column+52] = opt.wk52lo
         # _tradesheet[new_row_index, last_column+53] = opt.wk52lodate
         # _tradesheet[new_row_index, last_column+54] = opt.imp_volatility
         # _tradesheet[new_row_index, last_column+55] = opt.idelta
         # _tradesheet[new_row_index, last_column+56] = opt.igamma
         # _tradesheet[new_row_index, last_column+57] = opt.itheta
         # _tradesheet[new_row_index, last_column+58] = opt.ivega
         # _tradesheet[new_row_index, last_column+59] = opt.irho
         # _tradesheet[new_row_index, last_column+60] = opt.openinterest
         # last_column = last_column + 61

    return

# def find_profitable_trades( _tradesheet, _worksheet, _data, _bid, _ask):
#     covered_call( _tradesheet, _worksheet, _data, _bid, _ask )
#     iron_condor( _tradesheet, _worksheet, _data, _bid, _ask )
#     long_box( _tradesheet, _worksheet, _data, _bid, _ask )
#     short_box( _tradesheet, _worksheet, _data, _bid, _ask )
#     return
